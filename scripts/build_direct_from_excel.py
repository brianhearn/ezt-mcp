#!/usr/bin/env python3
"""Build Direct Build request assignments from an Excel ZIP-to-territory sheet."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_PATH_COLUMNS = ["Divison", "Region", "Territory"]
DEFAULT_PART_COLUMN = "Postal Code"


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", help="Path to .xlsx file")
    parser.add_argument("--sheet", default="ZIP2Terr")
    parser.add_argument("--part-column", default=DEFAULT_PART_COLUMN)
    parser.add_argument("--path-column", action="append", dest="path_columns")
    parser.add_argument("--part-layer", default="us_zips")
    parser.add_argument("--tal-label", default="PowerBI Demo ZIP2Terr")
    parser.add_argument("--tal-id", default="tal-powerbi-demo-zip2terr")
    parser.add_argument("--duplicate-policy", choices=["error", "first", "last"], default="error")
    parser.add_argument("--output", help="Write Direct Build request JSON to this path")
    parser.add_argument("--summary", help="Write summary JSON to this path")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    path_columns = args.path_columns or DEFAULT_PATH_COLUMNS
    result = build_request(
        Path(args.excel),
        sheet_name=args.sheet,
        part_column=args.part_column,
        path_columns=path_columns,
        part_layer=args.part_layer,
        tal_label=args.tal_label,
        tal_id=args.tal_id,
        duplicate_policy=args.duplicate_policy,
    )
    if args.output:
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)
        Path(args.output).write_text(json.dumps(result["request"], indent=2), encoding="utf-8")
    if args.summary:
        Path(args.summary).parent.mkdir(parents=True, exist_ok=True)
        Path(args.summary).write_text(json.dumps(result["summary"], indent=2), encoding="utf-8")
    print(json.dumps(result["summary"], indent=2))
    return 0


def build_request(
    excel_path: Path,
    *,
    sheet_name: str,
    part_column: str,
    path_columns: list[str],
    part_layer: str,
    tal_label: str,
    tal_id: str,
    duplicate_policy: str,
) -> dict[str, Any]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Sheet {sheet_name!r} not found. Available sheets: {workbook.sheetnames}")
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise SystemExit(f"Sheet {sheet_name!r} is empty") from exc
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    indexes = {header: idx for idx, header in enumerate(headers) if header}
    missing = [column for column in [part_column, *path_columns] if column not in indexes]
    if missing:
        raise SystemExit(f"Missing columns: {missing}. Headers: {headers}")

    assignments_by_part: dict[str, dict[str, Any]] = {}
    duplicate_conflicts: list[dict[str, Any]] = []
    exact_duplicates = 0
    row_count = 0
    invalid_rows: list[dict[str, Any]] = []
    territory_counts: Counter[str] = Counter()
    path_counts: Counter[tuple[str, ...]] = Counter()

    for excel_row_number, row in enumerate(rows, start=2):
        if not row or not any(value is not None for value in row):
            continue
        row_count += 1
        part_id = clean_zip(row[indexes[part_column]])
        territory_path = tuple(clean_label(row[indexes[column]]) for column in path_columns)
        if not part_id or not all(territory_path):
            invalid_rows.append(
                {
                    "row_number": excel_row_number,
                    "part_id": part_id,
                    "territory_path": list(territory_path),
                }
            )
            continue
        assignment = {"part_id": part_id, "territory_path": list(territory_path)}
        previous = assignments_by_part.get(part_id)
        if previous is not None:
            previous_path = tuple(previous["territory_path"])
            if previous_path == territory_path:
                exact_duplicates += 1
                continue
            duplicate_conflicts.append(
                {
                    "part_id": part_id,
                    "previous_territory_path": list(previous_path),
                    "conflicting_territory_path": list(territory_path),
                    "row_number": excel_row_number,
                }
            )
            if duplicate_policy == "error":
                continue
            if duplicate_policy == "first":
                continue
        assignments_by_part[part_id] = assignment
        territory_counts[territory_path[-1]] += 1
        path_counts[territory_path] += 1

    if duplicate_conflicts and duplicate_policy == "error":
        # Still emit a summary so callers can inspect the conflicts.
        assignments: list[dict[str, Any]] = []
    else:
        assignments = [assignments_by_part[key] for key in sorted(assignments_by_part)]

    request = {
        "part_layer": part_layer,
        "tal_label": tal_label,
        "tal_id": tal_id,
        "assignments": assignments,
        "source": {
            "type": "excel_sheet",
            "path": str(excel_path),
            "sheet": sheet_name,
            "part_column": part_column,
            "path_columns": path_columns,
            "duplicate_policy": duplicate_policy,
        },
    }
    summary = {
        "ok": not invalid_rows and (not duplicate_conflicts or duplicate_policy != "error"),
        "excel": str(excel_path),
        "sheet": sheet_name,
        "headers": headers,
        "row_count": row_count,
        "assignment_count": len(assignments),
        "unique_part_count": len(assignments_by_part),
        "invalid_row_count": len(invalid_rows),
        "exact_duplicate_count": exact_duplicates,
        "duplicate_conflict_count": len(duplicate_conflicts),
        "duplicate_policy": duplicate_policy,
        "path_depth": len(path_columns),
        "territory_count": len(territory_counts),
        "rollup_path_count": len({path[:-1] for path in path_counts}),
        "top_territories": territory_counts.most_common(10),
        "duplicate_conflicts": duplicate_conflicts[:50],
        "invalid_rows": invalid_rows[:50],
        "request_output_safe": bool(assignments),
    }
    return {"request": request, "summary": summary}


def clean_zip(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        raw = str(int(value))
    elif isinstance(value, int):
        raw = str(value)
    else:
        raw = str(value).strip()
        if raw.endswith(".0") and raw[:-2].isdigit():
            raw = raw[:-2]
    digits = re.sub(r"\D", "", raw)
    if not digits:
        return ""
    return digits.zfill(5)[:5]


def clean_label(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
